# export_summary.py
import os
import re
import argparse
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# 类别中英文映射字典
CLASS_TRANSLATION = {
    "Healthy": "健康",
    "Potentially unhealthy": "潜在不健康",
    "Unhealthy": "不健康",
    "Uncracked": "无裂缝",
    "Crack": "有裂缝"
}

def parse_log_file(log_path):
    """解析单个 log.txt 文件并提取数据项"""
    with open(log_path, 'r', encoding='utf-8') as f:
        content = f.read()

    # 清理特殊空白字符（如全角空格 \u3000 和不间断空格 \xa0）
    content = content.replace('\xa0', ' ').replace('\u3000', ' ')

    data = {}
    
    # 0. 实验元信息（仅保留日期）
    normalized_path = log_path.replace('\\', '/')
    path_parts = normalized_path.split('/')
    if len(path_parts) >= 4:
        data["日期\nDate"] = path_parts[-3]

    # 1. 提取配置与超参数 (增加兼容性匹配)
    dataset_m = re.search(r"数据集名称\s*\(Dataset Name\)\s*[:：]\s*([^\n]+)", content)
    sub_type_m = re.search(r"数据集子类型\s*\(Sub Type\)\s*[:：]\s*([^\n]+)", content)
    
    dataset_name = dataset_m.group(1).strip() if dataset_m else ""
    sub_type = sub_type_m.group(1).strip() if sub_type_m else ""
    
    if sub_type and sub_type.upper() != "NONE":
        data["数据集\nDataset"] = f"{dataset_name} ({sub_type})"
    else:
        data["数据集\nDataset"] = dataset_name

    model_type_m = re.search(r"模型名称\s*\(Model Type\)\s*[:：]\s*([^\n]+)", content) or re.search(r"模型名称\s*\(Model\)\s*[:：]\s*([^\n]+)", content)
    data["模型名称\nModel"] = model_type_m.group(1).strip() if model_type_m else ""

    model_id_m = re.search(r"模型ID\s*\(Model ID\)\s*[:：]\s*([^\n]+)", content)
    data["模型ID\nModel ID"] = model_id_m.group(1).strip() if model_id_m else ""

    lr_m = re.search(r"学习率\s*\(Learning Rate\)\s*[:：]\s*([^\n]+)", content)
    if lr_m:
        try:
            data["学习率\nLR"] = float(lr_m.group(1).strip())
        except ValueError:
            data["学习率\nLR"] = lr_m.group(1).strip()

    bs_m = re.search(r"批次大小\s*\(Batch Size\)\s*[:：]\s*(\d+)", content)
    data["批次大小\nBatch Size"] = int(bs_m.group(1).strip()) if bs_m else None

    wd_m = re.search(r"权重衰减\s*\(Weight Decay\)\s*[:：]\s*([^\n]+)", content)
    if wd_m:
        try:
            data["权重衰减\nWeight Decay"] = float(wd_m.group(1).strip())
        except ValueError:
            data["权重衰减\nWeight Decay"] = wd_m.group(1).strip()

    # 2. 样本数量统计
    samples_m = re.search(r"训练集样本数:\s*(\d+)\s*\|\s*验证集样本数:\s*(\d+)\s*\|\s*测试集样本数:\s*(\d+)", content)
    if samples_m:
        data["训练集样本\nTrain Size"] = int(samples_m.group(1))
        data["验证集样本\nVal Size"] = int(samples_m.group(2))
        data["测试集样本\nTest Size"] = int(samples_m.group(3))
        num_test_samples = int(samples_m.group(3))
    else:
        test_samples = re.search(r"测试集样本数:\s*(\d+)", content)
        num_test_samples = int(test_samples.group(1)) if test_samples else None
        if num_test_samples:
            data["测试集样本\nTest Size"] = num_test_samples

    # 3. 训练指标与轮数
    train_epochs = re.search(r"实际训练总轮数:\s*(\d+)/(\d+)", content)
    if train_epochs:
        data["训练轮数\nEpochs"] = f"{train_epochs.group(1)}/{train_epochs.group(2)}"

    best_epoch = re.search(r"最佳模型出现轮数:\s*Epoch\s*(\d+)", content)
    data["最佳轮次\nBest Epoch"] = int(best_epoch.group(1)) if best_epoch else None

    best_val_loss = re.search(r"Val Loss\s*:\s*([\d\.]+)\s*\|\s*Val Acc\s*:\s*([\d\.]+)", content)
    if best_val_loss:
        data["验证集准确率\nVal Acc [Best]"] = float(best_val_loss.group(2))
        data["验证集损失\nVal Loss [Best]"] = float(best_val_loss.group(1))

    best_train_loss = re.search(r"Train Loss:\s*([\d\.]+)\s*\|\s*Train Acc:\s*([\d\.]+)", content)
    if best_train_loss:
        data["最佳轮训练准确率\nTrain Acc [Val Best]"] = float(best_train_loss.group(2))
        data["最佳轮训练损失\nTrain Loss [Val Best]"] = float(best_train_loss.group(1))

    # 时间与延迟
    avg_train_time = re.search(r"每轮平均耗时:\s*([\d\.]+)s", content)
    data["单轮耗时\nEpoch Time (s)"] = float(avg_train_time.group(1)) if avg_train_time else None

    # 4. 测试集评估指标
    test_acc = re.search(r"测试集准确率 \(Overall Accuracy\):\s*([\d\.]+)%", content)
    data["测试集准确率\nTest Acc (%)"] = float(test_acc.group(1)) if test_acc else None

    test_time = re.search(r"测试集推理评估总耗时 \(Test Total Time\):\s*([\d\.]+)s", content)
    if test_time:
        total_time_s = float(test_time.group(1))
        data["测试总耗时\nTest Time (s)"] = total_time_s
        if num_test_samples:
            data["推理延迟\nLatency (ms/img)"] = round((total_time_s / num_test_samples) * 1000, 2)

    # 5. 全局 Average 指标
    macro_avg = re.search(r"macro avg\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)\s+(\d+)", content)
    if macro_avg:
        data["宏精准度\nMacro Prec"] = float(macro_avg.group(1))
        data["宏召回率\nMacro Rec"] = float(macro_avg.group(2))
        data["宏F1\nMacro F1"] = float(macro_avg.group(3))

    weighted_avg = re.search(r"weighted avg\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)\s+(\d+)", content)
    if weighted_avg:
        data["加权精准度\nWeighted Prec"] = float(weighted_avg.group(1))
        data["加权召回率\nWeighted Rec"] = float(weighted_avg.group(2))
        data["加权F1\nWeighted F1"] = float(weighted_avg.group(3))

    # 6. 解析分类报告，按类别提取 F1
    report_block_m = re.search(r"===+\s*最终评估结果\s*===+([\s\S]+?)==========================================", content)
    if report_block_m:
        report_block = report_block_m.group(1)
        lines = report_block.split('\n')
        for line in lines:
            line_str = line.strip()
            match = re.match(r"^([A-Za-z0-9_\-\s]+?)\s+([\d\.]+)\s+([\d\.]+)\s+([\d\.]+)\s+(\d+)$", line_str)
            if match:
                cls_name = match.group(1).strip()
                f1_val = float(match.group(4))
                
                if cls_name.lower() in ["accuracy", "macro avg", "weighted avg"]:
                    continue
                    
                zh_name = CLASS_TRANSLATION.get(cls_name, cls_name)
                col_label = f"{zh_name}(F1)\n{cls_name} F1"
                data[col_label] = f1_val

    return data


def main():
    parser = argparse.ArgumentParser(description="自动汇总模型评估结果至双语美化 Excel")
    parser.add_argument("--dataset", type=str, default="SDNET2018_D", help="目标数据集输出子目录名称")
    args = parser.parse_args()

    dataset_dir = os.path.join("outputs", args.dataset)
    if not os.path.exists(dataset_dir):
        print(f"[错误] 数据集目录不存在: {dataset_dir}")
        return

    log_files = []
    for root, _, files in os.walk(dataset_dir):
        for file in files:
            if file.endswith("_log.txt") or file == "train_log.txt":
                log_files.append(os.path.join(root, file))

    if not log_files:
        print(f"[提示] 未在 {dataset_dir} 路径下查找到任何日志文件！")
        return

    print(f"找到 {len(log_files)} 个日志文件，解析汇总中...")
    
    all_data = []
    for log_path in log_files:
        try:
            parsed_data = parse_log_file(log_path)
            if parsed_data:
                all_data.append(parsed_data)
        except Exception as e:
            print(f"[警告] 解析日志失败 {log_path}: {e}")

    df = pd.DataFrame(all_data)

    preferred_order = [
        "日期\nDate", "数据集\nDataset",
        "模型名称\nModel", "模型ID\nModel ID", 
        "训练集样本\nTrain Size", "验证集样本\nVal Size", "测试集样本\nTest Size",
        "学习率\nLR", "批次大小\nBatch Size", "权重衰减\nWeight Decay",
        "测试集准确率\nTest Acc (%)", 
        "最佳轮次\nBest Epoch", "训练轮数\nEpochs",
        "验证集准确率\nVal Acc [Best]", "验证集损失\nVal Loss [Best]",
        "最佳轮训练准确率\nTrain Acc [Val Best]", "最佳轮训练损失\nTrain Loss [Val Best]",
        "推理延迟\nLatency (ms/img)", "测试总耗时\nTest Time (s)", "单轮耗时\nEpoch Time (s)",
        "宏F1\nMacro F1", "加权F1\nWeighted F1",
        "宏精准度\nMacro Prec", "宏召回率\nMacro Rec",
        "加权精准度\nWeighted Prec", "加权召回率\nWeighted Rec"
    ]

    existing_cols = [c for c in preferred_order if c in df.columns]
    remaining_cols = [c for c in df.columns if c not in existing_cols]
    df = df[existing_cols + remaining_cols]

    output_excel_path = os.path.join(dataset_dir, f"{args.dataset}_summary.xlsx")
    
    with pd.ExcelWriter(output_excel_path, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name="模型评估汇总", index=False)
        
        worksheet = writer.sheets["模型评估汇总"]
        
        header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        header_font = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
        data_font = Font(name="Microsoft YaHei", size=9.5)
        
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
        
        worksheet.row_dimensions[1].height = 38
        for col_idx, col in enumerate(worksheet.columns, start=1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            lines = str(cell.value).split('\n')
            max_header_line_len = max(len(line) for line in lines) if lines else 0
            
            max_data_len = 0
            for row_idx in range(2, len(df) + 2):
                data_cell = worksheet.cell(row=row_idx, column=col_idx)
                data_cell.font = data_font
                data_cell.border = thin_border
                data_cell.alignment = Alignment(horizontal="center", vertical="center")
                val_str = str(data_cell.value or '')
                if len(val_str) > max_data_len:
                    max_data_len = len(val_str)
            
            col_letter = cell.column_letter
            calc_width = max(max_header_line_len + 5, max_data_len + 4, 14)
            worksheet.column_dimensions[col_letter].width = min(calc_width, 35)

        for row_idx in range(2, len(df) + 2):
            worksheet.row_dimensions[row_idx].height = 22

    print(f"\n[导出成功] 表格字段与样本数统计更新完成，保存至: {output_excel_path}")


if __name__ == "__main__":
    main()